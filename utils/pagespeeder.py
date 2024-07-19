from datetime import datetime
import openpyxl
import requests
import pandas as pd
from dotenv import load_dotenv
import os
# from lighthouse import LighthouseCI

load_dotenv() 


API_KEY = os.getenv("PAGESPEED_API_KEY")
HEADERS = {
    #'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
    'Content-Type': 'application/json',
    'Accept': 'application/json',
    #'Authorization': f'Bearer {API_KEY}'
}

try:
    urls = pd.read_csv('./urls.csv')['url'].tolist()
except FileNotFoundError:
    print("urls.csv file not found. Please make sure it's in the same directory as main.py")
    exit()
except pd.errors.EmptyDataError:
    print("urls.csv file is empty. Please make sure it's not empty, and it starts with a header (url).")
    exit()
except pd.errors.ParserError as e:
    print(f"Error parsing urls.csv: {e}")
    exit()



def get_page_speed_data(url, category='performance', strategy='mobile'):
    """Fetches Core Web Vitals, performance, accessibility, and SEO data for a given URL."""
    try:
        endpoint = f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={url}&category={category}&strategy={strategy}'
        response = requests.get(endpoint, headers=HEADERS)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f'Error: {e}')
        return None
        print(f'Failed to fetch data for {url}. : {e}')
        return None
    


# check_api_key = lambda: (requests.get('https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url=https://dinovix.com')).status_code == 200

def check_pagespeed():
    """Checks the PageSpeed result for the given URLs."""
    
    print('Checking URLs: \n')

    data = []

    for url in urls:
        print(f'\nChecking CWV & perfomance for {url}...')
        try:
            result = get_page_speed_data(url)
        except Exception as e:
            print(f'Failed to fetch data for {url}.')
            print(f'Error: {e}')
            continue

        if result is None:
            continue
        #TODO: check if metric exist in loding experiance and get inp
        if 'originLoadingExperience' in result and 'metrics' in result['originLoadingExperience']:
            try:
                inp = result['originLoadingExperience']['metrics']['INTERACTION_TO_NEXT_PAINT']['percentile']
                fid = result['originLoadingExperience']['metrics']['FIRST_INPUT_DELAY_MS']['percentile']
                ttfb = result['originLoadingExperience']['metrics']['EXPERIMENTAL_TIME_TO_FIRST_BYTE']['percentile']
            except KeyError as e:
                print(f'Failed to fetch INP, FID, or TTFB for {url}.')
                # print(f'Error: {e}')
                inp = "N/A"
                fid = "N/A"
                ttfb = "N/A"
                


        if 'lighthouseResult' in result:
            try:
                lcp = result['lighthouseResult']['audits']['largest-contentful-paint']['displayValue']
                fcp = result.get('lighthouseResult', {}).get('audits', {}).get('first-contentful-paint', {}).get('displayValue', '')
                cls = result.get('lighthouseResult', {}).get('audits', {}).get('cumulative-layout-shift', {}).get('displayValue', '')
                performance_score = f"{str(result['lighthouseResult']['categories']['performance']['score']*100)}%"
                
                if 'accessibility' in result['lighthouseResult']['categories']:
                    accessibility_score = f"{result['lighthouseResult']['categories']['accessibility']['score']*100}%"
                else:
                    result = get_page_speed_data(url, "accessibility")
                    accessibility_score = f"{result['lighthouseResult']['categories']['accessibility']['score']*100}%"
                
                if 'best-practices' in result['lighthouseResult']['categories']:
                    best_practices = f"{result['lighthouseResult']['categories']['best-practices']['score']*100}%"
                else:
                    result = get_page_speed_data(url, "best-practices")
                    best_practices = f"{result['lighthouseResult']['categories']['best-practices']['score']*100}%"
                if 'seo' in result['lighthouseResult']['categories']:
                    seo_score = f"{result['lighthouseResult']['categories']['seo']['score']*100}%"
                else:
                    result = get_page_speed_data(url, "seo")
                    seo_score = f"{result['lighthouseResult']['categories']['seo']['score']*100}%"

                # get overall loading experience message
                if 'overall_category' in result['loadingExperience']: 
                    overall_message = result['loadingExperience']['overall_category']
                else:
                    overall_message = ""
            except KeyError as e:
                print(f'Failed to fetch LCP, FCP, CLS, Performance, Accessibility, SEO or Overall Loading Experience for {url}.')
                print(f'Error: {e}')
                lcp = "N/A"
                fcp = "N/A"
                cls = "N/A"
                performance_score = "N/A"
                accessibility_score = "N/A"
                seo_score = "N/A"
                best_practices = "N/A"
                overall_message = "N/A"

            data.append({
                'Website URL': url,
                'LCP score': lcp,
                'FCP score': fcp,
                'CLS score': cls,
                'FID in ms': fid,
                'INP in ms': inp,
                'TTFB in ms': ttfb,
                'Performance': performance_score,
                'Accessibility': accessibility_score,
                'SEO score': seo_score,
                'Best Bractices': best_practices,
                'Overall Loading Experience': overall_message
            })

            print(f'{url} - LCP: {lcp}, INP: {inp},  FCP: {fcp}, CLS: {cls}, Performance: {performance_score}, Accessibility: {accessibility_score}, SEO: {seo_score} \n')
            emoji = ''
            if overall_message.capitalize() == 'Slow' or overall_message == None:
                emoji = '👎'
            elif overall_message.capitalize()  == 'Average':
                emoji = '⚠️' 
            elif overall_message.capitalize()  == 'Fast':
                emoji = '👍'
            print(f'CWV & perfomance for {url} checked, experience: {overall_message or None} {emoji}.\n\n')
        else:
            print(f'CWV & perfomance for {url} not found.\n\n')
            continue

    if data:
        try:
            now = datetime.now().strftime("%Y-%m-%d_%Hh%Mm.xlsx")
            filename = f'cwv_report_{now}'
            write_to_excel_file_and_format(data, filename)
            print(f'Report saved to {filename}')
        except Exception as e:
            print(f'Failed to save report to {filename}.')
            print(f'Error: {e}')

def write_to_excel_file_and_format(data, filename):
    """Write data to an Excel file and format the cells."""
    df = pd.DataFrame(data)
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    df.to_excel(filename, index=False)
    print(f'Data written to {filename}.')

    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for column_cells in ws.iter_cols(max_row=1):
        max_length = 10
        for cell in column_cells:
            if cell.value is not None:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
        ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = 35 if column_cells[0].column == 1 else (max_length + 3)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].height = 25

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                
                if cell is not row[0]:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

                if cell is row[1]: #LCP
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color= "90EE90" if float(cell.value.rstrip(' s')) <= 2.5 else "FFCC00" if float(cell.value.rstrip('s')) <= 4 else "FFCCCC")
                if cell is row[2]: #FCP
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color=("90EE90" if float(cell.value.rstrip(' s')) <= 1.8 else "FFCC00" if float(cell.value.rstrip('s')) <= 3 else "FFCCCC"))
                if cell is row[3]: #CLS
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color=("90EE90" if float(cell.value.rstrip(' s')) <= 0.1 else "FFCC00" if float(cell.value.rstrip('s')) <= 0.25 else "FFCCCC"))
                if cell is row[7]: #Performance
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color=("90EE90" if float(cell.value.rstrip('%')) >= 90 else "FFCC00" if float(cell.value.rstrip('%'))  >= 50 else "FFCCCC"))
                if cell is row[8]: #Acc
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color=("90EE90" if float(cell.value.rstrip('%')) >= 90 else "FFCC00" if float(cell.value.rstrip('%'))  >= 50 else "FFCCCC"))
                if cell is row[9]: #Seo
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color=("90EE90" if float(cell.value.rstrip('%')) >= 90 else "FFCC00" if float(cell.value.rstrip('%'))  >= 50 else "FFCCCC"))
                if cell is row[10]: #Best pr.
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color=("90EE90" if float(cell.value.rstrip('%')) >= 90 else "FFCC00" if float(cell.value.rstrip('%'))  >= 50 else "FFCCCC"))

                # Overall Loading Experience cells should be coloFFCCCC based on their value
                if cell.value == 'FAST':
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='00FF00')
                elif cell.value == 'AVERAGE':
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='FFFF00')
                elif cell.value == 'SLOW':
                    cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='FF0000') 
        
    wb.save(filename)
    wb.close()
